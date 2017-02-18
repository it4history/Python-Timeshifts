# -*- coding: utf-8 -*-

import xlrd, xlwt
from openpyxl import load_workbook, Workbook
#from difflib import Differ, get_close_matches
from difflib import get_close_matches
from threading import Thread

#from xlutils.filter import GlobReader,BaseFilter,DirectoryWriter,process
import os
import re                
recomp = re.compile(ur"""
\d+
|(бря|бре|июня|июля|июне|июле)
|(-|–|—)
""",re.VERBOSE)

myfile=['00_orig_00.xls', '00_orig_01.xls', '00_orig_02.xls', '00_orig_03.xls']
new_myfile=['00_out_00.xls','00_out_01.xls','00_out_02.xls','00_out_03.xls']
mydir='C:/Users/drondin/Documents/'
#mydir='C:/Users/Im Nox/Documents/'

file_type = myfile[0].split(".")[-1]

shifts = [83,167,251,334,417,501,18,23,36,41,59,64,83,100,72,76,77,108]
exceptions = {"a0": u"а", "a1": u"А",
              "bil0": u"был", "bil1": u"Был", "bila0": u"была", "bila1": u"Была",
              "bili0": u"были", "bili1": u"Были",
              "v0": u"в", "v1": u"В", "vo": u"во","velik": u"велик","velika": u"велика",
              "g_dot": u"г.", "gde": u"где", "god0": u"год","god_dot": u"год.", "god1": u"Год",
              "godu0": u"году", "godu_dot": u"году.", "godu1": u"Году",
              "do0": u"до", "do1": u"До", "esli0": u"если", "esli1": u"Если",
              "i0": u"и", "i1": u"И", "iz0": u"из", "iz1": u"Из", "ili0": u"или", "ili1": u"Или",
              "k0": u"к", "k1": u"К", "kak0": u"как", "kak1": u"Как", "kotorogo": u"которого",
              "kotorii": u"который", "kotoraya": u"которая",
              "l_dot": u"л.", "li": u"ли", "libo0": u"либо", "libo1": u"Либо", "m_dot": u"м.", "mi": u"мы",
              "na0": u"на", "na1": u"На","nas0": u"нас", "nas1": u"Нас",
              "ne": u"не", "net": u"нет", "nih": u"них","n_dot_e_dot": u"н.э.",
              "o0": u"о", "o1": u"О", "on0": u"он", "on1": u"Он",
              "ona0": u"она", "ona1": u"Она","oni0": u"они", "oni1": u"Они", "ono0": u"оно", "ono1": u"Оно",
              "ot0": u"от", "ot1": u"От", "ob0": u"об", "ob1": u"Об",
              "pered0": u"перед", "pered1": u"Перед", "po0": u"по", "po1": u"По",
              "r_dot": u"р.",
              "s0": u"с", "s1": u"С", "stal": u"стал", "stalo0": u"стало", "stalo1": u"Стало",
              "takje0": u"также", "takje1": u"Также","to0": u"то", "to1": u"То", "toje0": u"тоже",
              "toje1": u"Тоже",
              "h_dot": u"х.","chto0": u"что", "chto1": u"Что", "chtobi1": u"Чтобы", "chtobi0": u"чтобы",
              "eto0": u"это", "eto1": u"Это", "etom0": u"этом", "etom1": u"Этом"}

'''def iter_rows(sheet, rng):
    for row in sheet.iter_rows(rng, row_offset=1):
        for cell in row:
            yield [cell.value for cell in row]'''

def Filter(beg, end, f1, f2, types):
    MTCH = 1
    row_number = 0

    if file_type == "xls":
        font0 = xlwt.Font()
        font0.bold = False
        style0 = xlwt.XFStyle()
        style0.font = font0
        wb = xlwt.Workbook()
        ws = wb.add_sheet('A Test Sheet')
        rb = xlrd.open_workbook(os.path.join(mydir, myfile[f1]), formatting_info=True)
        sheet = rb.sheet_by_index(0)
        years = [sheet.row_values(rownum)[0] for rownum in xrange(sheet.nrows)]
        years.reverse()
        l_years = len(years)
        
        for i in xrange(beg, end):
            if row_number > 65536:
                print "rows count > 65536 in thread {0}".format(f1+1)
                break
            print i
            if types == "B":
                row_v_i = sheet.row_values(rownum-i)[1]
                row_v_i_nn = recomp.sub('', row_v_i)
                for j in xrange(i, l_years):
                    abs_years = abs(years[i] - years[j])
                    if abs_years in shifts:
                        row_v_j = sheet.row_values(rownum-j)[1]
                        new = [abs_years, years[i], years[j], row_v_i, row_v_j]
                        m = []
                        for k in row_v_i_nn.split():
                            if not k in exceptions.values():
                                matches = get_close_matches(k, row_v_j.split())
                                m.extend(matches)
                        l_mtch = len(m)
                        if l_mtch >= MTCH:
                            new.append(l_mtch)
                            for l in xrange(0, len(new)):
                                ws.write(row_number, l, new[l], style0)
                            row_number += 1
            elif types == "C":
                row_v_i = sheet.row_values(rownum-i)[1]
                row_v_i2 = sheet.row_values(rownum-i)[2]
                for j in xrange(i, l_years):
                    abs_years = abs(years[i] - years[j])
                    if abs_years in shifts:
                        row_v_j = sheet.row_values(rownum-j)[1]
                        row_v_j2 = sheet.row_values(rownum-j)[2]
                        new = [abs_years, int(years[i]), int(years[j]), row_v_i, row_v_j, row_v_i2, row_v_j2]
                        if row_v_i2 == row_v_j2:
                            for l in xrange(0, len(new)):
                                ws.write(row_number, l, new[l], style0)
                            row_number += 1
            wb.save(os.path.join(mydir, new_myfile[f2]))
        
    elif file_type == "xlsx":
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title='A Test Sheet')
        rb = load_workbook(filename = os.path.join(mydir, myfile[f1]), read_only=True)
        sheet = rb[u'База']
        years = [int(row[0].value) for row in sheet.iter_rows('A1:A28937', row_offset=1)]
        rownum = sheet.max_row - 1
        years.reverse()
        l_years = len(years)
    
        for i in xrange(beg, end):
            if row_number > 65536:
                print "rows count > 65536 in thread {0}".format(f1+1)
                break
            print i
            if types == "B":
                row_v_i = sheet.cell(row=rownum-i, column=2).value
                row_v_i_nn = recomp.sub('', row_v_i)
                for j in xrange(i, l_years):
                    abs_years = abs(years[i] - years[j])
                    if abs_years in shifts:
                        row_v_j = sheet.cell(row=rownum-j, column=2).value
                        new = [abs_years, years[i], years[j], row_v_i, row_v_j]
                        m = []
                        for k in row_v_i_nn.split():
                            if not k in exceptions.values():
                                matches = get_close_matches(k, row_v_j.split())
                                m.extend(matches)
                        l_mtch = len(m)
                        if l_mtch >= MTCH:
                            new.append(l_mtch)
                            ws.append(new)
                            row_number += 1
            elif types == "C":
                row_v_i = sheet.cell(row=rownum-i, column=2).value
                row_v_i2 = sheet.cell(row=rownum-i, column=3).value
                for j in xrange(i, l_years):
                    abs_years = abs(years[i] - years[j])
                    if abs_years in shifts:
                        row_v_j = sheet.cell(row=rownum-j, column=2).value
                        row_v_j2 = sheet.cell(row=rownum-j, column=3).value
                        new = [abs_years, int(years[i]), int(years[j]), row_v_i, row_v_j, row_v_i2, row_v_j2]
                        if row_v_i2 == row_v_j2:
                            ws.append(new)
                            row_number += 1
            wb.save(filename = os.path.join(mydir, new_myfile[f2]))

t1 = Thread(target=Filter, args=(27000,28001,0,0,"C"))
t2 = Thread(target=Filter, args=(28001,28938,1,1,"C"))
#t3 = Thread(target=Filter, args=(12601,12901,2,2,"C"))
#t4 = Thread(target=Filter, args=(12901,13201,3,3,"C"))

t1.start()
t2.start()
#t3.start()
#t4.start()
t1.join()
t2.join()
#t3.join()
#t4.join()

'''
class ShiftsFilter(BaseFilter):

    goodlist = []
    
    def __init__(self,elist): 
        self.goodlist = goodlist
        self.wtw = 0
        self.wtc = 0
         

    def workbook(self, rdbook, wtbook_name): 
        self.next.workbook(rdbook, 'filtered_'+wtbook_name) 

    def row(self, rdrowx, wtrowx):
        pass

    def cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        value = self.rdsheet.cell(rdrowx,rdcolx).value
        if value in self.goodlist:
            self.wtc=self.wtc+1 
            self.next.row(rdrowx,wtrowx)
        else:
            return
        self.next.cell(rdrowx,rdcolx,self.wtc,wtcolx)
        
        
data = """somedata1
somedata2
somedata3
somedata4
somedata5
"""

goodlist = data.split("\n")

process(GlobReader(os.path.join(mydir,myfile)), ShiftsFilter(goodlist), DirectoryWriter(mydir))'''

